from openpyxl import load_workbook
from docxtpl import DocxTemplate, Listing

wb = load_workbook('26.Scratch编程课程信息汇总.xlsx')
ws = wb.active

doc = DocxTemplate('1.docx') #加载模板文件

data = tuple(ws.iter_rows(min_row=2, max_row=35, min_col=5, max_col=8))

filled_in_data = {
    'class_name': data[1][0].value,
    'class_purpose': Listing(data[1][2].value),
    'class_knowledge': Listing(data[1][3].value),
    'lecturor': '图图',

    'student_name' : '{{student_name}}',
    'communication' : '{{communication}}',
    'creation' : '{{creation}}',
    'co_operation' : '{{co_operation}}',
    'solvability' : '{{solvability}}',
    'thoughts' : '{{thoughts}}',
    'lecturer_comments' : '{{lecturer_comments}}',
    'year' : '{{year}}',
    'month' : '{{month}}',
    'day' : '{{day}}'
}

document_name = '02' + data[1][0].value + '.docx'

doc.render(filled_in_data)
doc.save(document_name)

print(document_name + '已生成')