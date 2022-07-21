import os, easyocr, datetime, re, tkinter as tk
from tkinter import filedialog
from openpyxl.workbook import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.worksheet.datavalidation import DataValidation


now = datetime.datetime.now()
month = now.month
new_file_name = '图图'+str(month)+'月报销表单.xlsx'

# 处理图片信息的函数
def get_transaction_information(info_list):
    pattern_quanitiy_type1 = re.compile('x+')
    pattern_quanitiy_type2 = re.compile('X+')
    pattern_price = re.compile('半+')
    pattern_time = re.compile('创建时间+')
    count_product = 0
    transcation_information = {}

    for item in info_list:
        match_quanitiy_type1 = pattern_quanitiy_type1.findall(item)
        match_quanitiy_type2 = pattern_quanitiy_type2.findall(item)
        match_price = pattern_price.findall(item)
        match_time = pattern_time.findall(item)
        if match_quanitiy_type1 or match_quanitiy_type2:
            count_product += 1
            # print(count_product)
            product_name = info_list[info_list.index(item)-1]
            product_quanitiy = item
            transcation_information[product_name] = product_quanitiy
            # print('商品数量',item)
        if match_price:
            title_index = info_list.index(item)-1
            # print(info_list.index(item))
            if item == '半':
                # 如果包含半，实际为人民币单位符号，前一元素为名称，后一元素为数量
                value_index = title_index+2
                transcation_information[info_list[title_index]] = info_list[value_index]
                # print(info_list[title_index],info_list[value_index])
            else:
                # 在部分元素中，如实付款，数值与半连接在一起，需要去除半符号
                value_index = title_index+1
                transcation_information[info_list[title_index]] = info_list[value_index].replace('半','')
                # print(info_list[title_index],info_list[value_index].replace('半',''))
        if match_time:
            transcation_time_stamp = info_list[info_list.index(item)+1]
            transcation_information[item] = transcation_time_stamp
            # print('创建时间', transcation_time_stamp)

    if count_product>1 or ('支付成功' in info_list):
        print('交易截图中有两个商品，请重新选择图片')
        return None
    else:
        print(transcation_information)
    return transcation_information

# 创建新的Excel表格
wb = Workbook()
ws = wb.active

# 设置对齐方式
align = Alignment(horizontal='center', vertical='center')
border = Border(left=Side(border_style='thin', color='000000'), right=Side(border_style='thin', color='000000'), top=Side(border_style='thin', color='000000'), bottom=Side(border_style='thin', color='000000'))

# 设置下拉菜单样式
dv_section = DataValidation(type="list", formula1='"科学启蒙,科学探究,工程实践,乐高,玛塔,程小奔,Scratch,Python,C++,假期营,校外幼儿园课,校外小学课,活动,教务行政用品,其他"', allow_blank=True)
dv_employee = DataValidation(type="list", formula1='"靳朴鈜,刘碧堃,王伟杰,杨敏,胡婷,徐敏,杨雪婷,李佳敏,巴图,张鹏,其他人"', allow_blank=True)
dv_payment_methods = DataValidation(type="list", formula1='"微信,支付宝,现金,银行卡,其他"', allow_blank=True)
ws.add_data_validation(dv_section)
ws.add_data_validation(dv_employee)
ws.add_data_validation(dv_payment_methods)

# 设置标题与表头
a1 = ws['A1']
a1.value = '运营成本记录'
a1.font = Font(name=u'黑体', size=12, bold=True)
a1.alignment = align
a1.border = border

title = ['日期', '款项', '用途', '单价', '数量', '运费', '优惠', '税额', '总金额', '付款人', '付款方式', '附件标号', '校验']
size = [16.67, 28, 24, 14.83, 14.83, 14.83, 14.83, 14.83, 14.83, 14.83, 14.83, 14.83, 14.83]
x = 0
for item in title:
    x += 1
    ws.cell(row=2, column=x, value=item)
    ws.cell(row=2, column=x).font = Font(name=u'黑体', size=12)
    ws.cell(row=2, column=x).alignment = align
    ws.cell(row=2, column=x).border = border
    
for x in range(ord('A'), ord('N')):
    ws.column_dimensions['{}'.format(chr(x))].width = size[x-65]
for x in range(0,66):
    ws.row_dimensions[x].height = 25
# 合并第一行
ws.merge_cells('A1:M1')

# 添加数据
def add_info(row,col,value):
    ws.cell(row=row, column=col, value=value)
    ws.cell(row=row, column=col).font = Font(name=u'黑体', size=12)
    ws.cell(row=row, column=col).alignment = align
    ws.cell(row=row, column=col).border = border

# 给所有单元格添加边框
for y in range(3,64):
    for x in range(1,14):
        add_info(y,x,'')

# 添加下拉菜单
dv_section.add('C1:C1048576')
dv_employee.add('J1:J1048576')
dv_payment_methods.add('K1:K1048576')

i65 = ws['I65']
i65.value = '=SUM(I3:I64)'
i65.font = Font(name=u'黑体', size=12, bold=True)
i65.alignment = align
i65.border = border
i65.fill=PatternFill('solid',fgColor='FFFF00')
# 完成表格的设计
    
# 选择交易截图
root = tk.Tk()
root.withdraw()

f_path = filedialog.askopenfilenames()
# print('\n获取的文件地址：', f_path)
good_pic_list = []

reader = easyocr.Reader(['ch_sim','en'], gpu=False)

row = 3
for file in f_path:
    result = reader.readtext(file, detail = 0)
    print(result)
    a = get_transaction_information(result)
    if a == None:
        ws['L'+str(row)] = os.path.split(file)[1]
        row += 1
        continue
    else:
        # a.values()
        product_name = []
        for item in a.keys():
            product_name.append(item)
        product_name = product_name[0]
        print(product_name)

        # 处理商品单价与数量
        product_value_quanitiy = []
        for item in a.values():
            product_value_quanitiy.append(item)
        product_value_quanitiy = product_value_quanitiy[0:2]
        # print(product_value_quanitiy)

        # 处理交易时间信息，只保留年/月/日
        try:
            t = datetime.datetime.strptime(a['创建时间:'],'%Y-%m-%d%H:%M:%S')
            t = datetime.datetime.strftime(t, '%Y/%m/%d')
        except:
            t = ''

        # 判断是否包含运费信息
        if '运费(快递)' in a.keys():
            shipment = float(a['运费(快递)'])
        elif '运费' in a.keys():
            shipment = float(a['运费'])
        else:
            shipment = ''

        # 去掉数量前的x
        try:
            if 'x' in product_value_quanitiy[1]:    
                quantity = int(product_value_quanitiy[1].strip('x'))
            else:
                quantity = int(product_value_quanitiy[1].strip('X'))
        except:
            quantity = ''

        # 填写表格
        try:
            equation = '='+'D'+str(row)+'*'+'E'+str(row)+'+'+'F'+str(row)+'-'+'G'+str(row)+'+'+'H'+str(row)
            ws['A'+str(row)] = t
            ws['B'+str(row)] = product_name
            ws['D'+str(row)] = float(product_value_quanitiy[0])
            ws['E'+str(row)] = quantity
            ws['F'+str(row)] = shipment
            ws['H'+str(row)] = 0
            ws['I'+str(row)] = equation
            ws['J'+str(row)] = '图图'
            ws['K'+str(row)] = '银行卡'
            ws['L'+str(row)] = os.path.split(file)[1]
            
            good_pic_list.append(file) 
        except:
            pass

        # 换行
        row += 1

location = os.path.split(f_path[0])[0]
wb.save(os.path.join(location, new_file_name))


print('无法处理的图片：', end='')
bad_pic_list = (set(f_path)-set(good_pic_list))
for pic in bad_pic_list:
    print(os.path.split(pic)[1], end='  ')