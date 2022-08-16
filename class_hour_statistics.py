# 课时统计

from ics import Calendar
from dateutil.relativedelta import relativedelta
from openpyxl import load_workbook
import requests, datetime, os

url = 'https://outlook.live.com/owa/calendar/3d1bd554-c7f0-49e2-996b-2cf4a6b39fa7/685e27f6-3554-4af1-b07c-4d785a347cc5/cid-28689F54F556F71F/calendar.ics'
c = Calendar(requests.get(url).text)
# c = Calendar(r.text)  # 使用本地ics文件访问

now = datetime.datetime.now()
year = now.year
month = now.month
start_date = datetime.date(year,month,1)
end_date = start_date + relativedelta(months=1)

new_file_name = '图图'+str(year)+'年'+str(month)+'月课时统计.xlsx'

e = list(c.timeline)    # 日历中的全部事件


class_in_total = []
for calender_event in e:    # 将事件添加到class_in_total中
    current_class = []
    if calender_event.begin.date() >= start_date and calender_event.begin.date() < end_date: # 筛选指定日期内的事件
        try:
            c1ass = calender_event.name.split(' ')[0]
            student = len(calender_event.name.split(' ')[1:])
            class_date = calender_event.begin.format('YYYY-MM-DD HH:mm').split(' ')[0]
            class_time = calender_event.begin.format('YYYY-MM-DD HH:mm').split(' ')[1]
            current_class.append(c1ass)
            current_class.append(class_date)
            current_class.append(class_time)
            current_class.append(student)
        except IndexError:
            pass

        class_in_total.append(current_class)
class_in_total.sort()
print('日历中共有'+str(len(class_in_total))+'项日程')

wb = load_workbook(filename = 'assest/教师课时统计表模板.xlsx')
ws = wb.active

cols = {'启蒙':2, '玛塔':5, '程小奔':5, '探究':8, '工程':11, 'Scratch':14, 'Python':14, '南开':17, '普林斯顿':17, '博苑澳森':17, '考古营':24, '医学营':24}

def insert_course_information(calender_list_item, course):  # 将内容添加至excel中
    time_value = datetime.datetime.strptime(calender_list_item[1],'%Y-%m-%d')   # 整理时间格式
    time_value = time_value.strftime('%m月%d日')
    if calender_list_item[0] == course:
        row = 5     # 固定在第5行开始
        if course not in cols.keys():   # 跳过日历中的其他事项
            print('未添加 ',time_value,course)
        else:
            col = cols[course]  # 从cols中获取课程所在列

            while ws.cell(row=row, column=col).value != None :  # 判断该单元格是否为空
                # print(str(row) + ' ' +str(col) + ' is not empty')
                row += 1
                if ws.cell(row=row, column=col).value == None:
                    break

            if course in ('南开','普林斯顿','博苑澳森'):    # 课程格式不同
                ws.cell(row=row, column=col, value=course)
                ws.cell(row=row, column=col+1, value=time_value)
                ws.cell(row=row, column=col+2, value=calender_list_item[2])
                ws.cell(row=row, column=col+3, value=calender_list_item[3])
                print('已添加',time_value,course,calender_list_item[3],'人')
                row += 1
                wb.save(filename='教师课时统计表模板.xlsx')
            else:
                ws.cell(row=row, column=col, value=time_value)
                ws.cell(row=row, column=col+1, value=calender_list_item[2])
                ws.cell(row=row, column=col+2, value=calender_list_item[3])
                print('已添加',time_value,course,calender_list_item[3],'人')
                row += 1
                wb.save(filename='教师课时统计表模板.xlsx')

for item in class_in_total:
    insert_course_information(item, item[0])
print('已导入'+str(len(class_in_total))+'项日程')

os.rename('教师课时统计表模板.xlsx', new_file_name)